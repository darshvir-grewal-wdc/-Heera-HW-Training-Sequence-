import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_excel_sheet(file_path, data):
    wb = Workbook()
    ws = wb.active

    # Applying formatting to the header row
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow color fill
    header_border = Border(bottom=Side(border_style="thick"))

    # Applying formatting to alternate rows
    alternate_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")  # Alternate color fill

    # Applying formatting to all cells
    cell_border = Border(left=Side(border_style="thin"),
                         right=Side(border_style="thin"),
                         top=Side(border_style="thin"),
                         bottom=Side(border_style="thin"))
    cell_alignment = Alignment(horizontal='center', vertical='center')
    first_column_alignment = Alignment(horizontal='center', vertical='center')
    first_column_font = Font(bold=True)

    # Dictionary to store the merge ranges
    merge_ranges = {}

    for row_idx, row_data in enumerate(data, 1):
        ws.append(row_data)
        if row_idx == 1:
            # Header row formatting
            for cell in ws[row_idx]:
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = header_fill
                cell.border = header_border
        elif row_idx % 2 == 0:
            # Alternate row formatting
            for cell in ws[row_idx]:
                cell.fill = alternate_fill
                cell.border = cell_border
        else:
            # Other row formatting
            for cell in ws[row_idx]:
                cell.border = cell_border

        # Formatting all cells
        for cell in ws[row_idx]:
            cell.alignment = cell_alignment

        # Formatting the first column
        first_cell = ws.cell(row=row_idx, column=1)
        first_cell.alignment = first_column_alignment
        first_cell.font = first_column_font

        # Merging cells in the first column
        cell_value = first_cell.value
        if cell_value in merge_ranges:
            merge_ranges[cell_value][1] = row_idx
        else:
            merge_ranges[cell_value] = [row_idx, row_idx]

    # Applying cell merging in the first column
    for merge_range in merge_ranges.values():
        start_row, end_row = merge_range
        ws.merge_cells(start_row=start_row, end_row=end_row, start_column=1, end_column=1)

    # Auto-adjusting column widths
    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

    wb.save(file_path)

def read_csv_as_list_of_lists(file_path):
    data = []
    with open(file_path, 'r') as csv_file:
        reader = csv.reader(csv_file)
        for row in reader:
            data.append(row)
    return data

def remove_empty_lists(data):
    return [row for row in data if row]

file_path = 'HW_Chaz.csv'

csv_data = remove_empty_lists(read_csv_as_list_of_lists(file_path))
header = ["TM SPeed","Chip & Die","Read Setup","Read Hold","Read Setup Minimum Passing Time (in ns)","Read Hold Minimum Passing Time (in ns)","ASIC Spec (in ns)","Expected Scope Vw (in ns)","Vw Margin w.r.t. ASIC Spec (in ns)","Write Setup","Write Hold","Write Setup Minimum Passing Time (in ns)","Write Hold Minimum Passing Time (in ns)","NAND Spec (in ns)","Expected Scope Vw (in ns)","Vw Margin w.r.t. NAND Spec (in ns)"]

#Time period in nano seconds
tab = [1.5,1.88,2.5,5.0]
ASIC_Spec = 0.15
NAND_Spec = 0.30
count = 0
tab_counter=-1

while count < len(csv_data):
    if (count % (len(csv_data)/4) == 0):
        tab_counter += 1
    csv_data[count].insert(3,str(round((int(csv_data[count][1]))/256*tab[tab_counter],3)))
    csv_data[count].insert(6,str(round((int(csv_data[count][4]))/256*tab[tab_counter],3)))
    csv_data[count].insert(4,str(round((tab[tab_counter]-float(int(csv_data[count][2])/256*tab[tab_counter])),3)))
    csv_data[count].insert(8,str(round((tab[tab_counter]-float(int(csv_data[count][6])/256*tab[tab_counter])),3)))
    csv_data[count].insert(5,str(round(ASIC_Spec,3)))
    csv_data[count].insert(10,str(round(NAND_Spec,3)))
    csv_data[count].insert(6,str(round(((int(csv_data[count][2])-int(csv_data[count][1]))/256*tab[tab_counter]),3)))
    csv_data[count].insert(12,str(round(((int(csv_data[count][8])-int(csv_data[count][7]))/256*tab[tab_counter]),3)))
    csv_data[count].insert(7,str(round(float(csv_data[count][6])-2*float(csv_data[count][5]),3)))
    csv_data[count].insert(14,str(round(float(csv_data[count][13])-2*float(csv_data[count][12]),3)))
    count += 1

#print(csv_data)
count = 0

while count < len(csv_data)/4:
    csv_data[count].insert(0,"TM667")
    count += 1

while count < 2*len(csv_data)/4:
    csv_data[count].insert(0,"TM533")
    count += 1

while count < 3*len(csv_data)/4:
    csv_data[count].insert(0,"TM400")
    count += 1

while count < 4*len(csv_data)/4:
    csv_data[count].insert(0,"TM200")
    count += 1

csv_data.insert(0,header)
file_path = 'HW_Characterization.xlsx'
create_excel_sheet(file_path, csv_data)
print(f"Excel sheet created at {file_path}.")
